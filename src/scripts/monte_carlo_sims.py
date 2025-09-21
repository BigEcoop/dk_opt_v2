#!/usr/bin/env python3
# src/scripts/monte_carlo_sims.py

import json
import random
import statistics
import argparse
import pandas as pd
import nfl_data_py as nfl
import urllib.error
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy

# ── PROJECT ROOT & PATHS ───────────────────────────────────────────────────────
PROJECT_ROOT    = Path(__file__).resolve().parents[2]
PROJS_ROOT      = PROJECT_ROOT / "src" / "output" / "projections"
SCORING_JSON    = PROJECT_ROOT / "src" / "data" / "scoring.json"
SIMS_ROOT       = PROJECT_ROOT / "src" / "output" / "sims"
ESPN_DIR        = PROJECT_ROOT / "src" / "data" / "projection_data" / "espn"
VEGAS_DIR       = PROJECT_ROOT / "src" / "data" / "projection_data" / "vegas"
SITUATION_DIR   = PROJECT_ROOT / "src" / "data" / "projection_data" / "situational"
DEF_RANK_DIR    = PROJECT_ROOT / "src" / "data" / "projection_data" / "defensive_ranks"
TEAM_PROFILE    = PROJECT_ROOT / "src" / "data" / "projection_data" / "2025_team.json"

# ── TUNING CONSTANTS ────────────────────────────────────────────────────────────
DISPERSION_K    = 13
ESPN_WEIGHT     = 0.04
FG_LAMBDA       = 1.2
SCORE_DAMPEN    = 0.75
CLAMP_MULT      = 2.0

team_map = {"LA": "LAR", "LAR": "LAR", "FA": None}

OUTPUT_CATEGORIES = [
    "pass_attempts","pass_completions","passing_yds","passing_tds","interceptions",
    "rushing_attempts","rushing_yds","rushing_tds",
    "receiving_targets","receptions","receiving_yds","receiving_tds",
    "fumbles","punt_return_yards","punt_return_tds","kick_return_yards","kick_return_tds"
]
OFFENSE_POSITIONS = {"qb", "rb", "wr", "te"}

def sample_nb(mu: float, k: int = DISPERSION_K) -> int:
    if mu <= 0:
        return 0
    p = k / (k + mu)
    return np.random.negative_binomial(k, p)

def load_json(path: Path) -> Dict:
    return json.loads(path.read_text())

def write_json(path: Path, data: Dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2))

def write_csv(path: Path, header: List[str], rows: List[List]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        f.write(",".join(header) + "\n")
        for row in rows:
            f.write(",".join(str(x) for x in row) + "\n")

def prepare_players(baseline_path: Path, week: int) -> List[Dict]:
    j   = load_json(baseline_path)
    raw = j.get("players", [])
    try:
        year   = int(baseline_path.stem.split("_")[2])
        inj_df = nfl.import_injuries([year])
    except (urllib.error.HTTPError, ValueError):
        inj_df = pd.DataFrame(columns=["player_name","injury_status"])
    injured = set(
        inj_df[inj_df["injury_status"].isin({"questionable","out","doubtful"})][
            "player_name"
        ]
    )

    players: List[Dict] = []
    for rec in raw:
        name   = rec.get("name") or rec.get("player") or rec.get("Player")
        pid    = rec.get("id")   or rec.get("ID")
        raw_tm = rec.get("team") or rec.get("Team")
        team   = team_map.get(raw_tm, raw_tm)
        pos    = (rec.get("pos") or rec.get("POS") or rec.get("position") or "").lower()

        if not (name and pid and team) or team is None:
            continue
        if name in injured or pos not in OFFENSE_POSITIONS:
            continue

        stats_raw = rec.get("stats", {})
        stats: Dict[str, float] = {}
        for st, mu in stats_raw.items():
            if st in OUTPUT_CATEGORIES:
                try:
                    stats[st] = float(mu)
                except (ValueError, TypeError):
                    stats[st] = 0.0

        if stats:
            players.append({
                "player":   name,
                "id":       pid,
                "team":     team,
                "position": pos,
                "stats":    stats
            })

    return players

def score_statline(statline: Dict[str, float], scoring: Dict) -> float:
    o, y, b = scoring["offense"], scoring["yardage"], scoring["bonuses"]
    pts = 0.0
    pts += statline.get("passing_tds", 0)   * o["passing_td"]
    pts += statline.get("interceptions", 0) * o["interception"]
    pts += statline.get("passing_yds", 0)   * y["passing"]
    if statline.get("passing_yds", 0) >= 300:
        pts += b["passing_300_yds"]
    pts += statline.get("rushing_tds", 0)   * o["rushing_td"]
    pts += statline.get("rushing_yds", 0)   * y["rushing"]
    if statline.get("rushing_yds", 0) >= 100:
        pts += b["rushing_100_yds"]
    pts += statline.get("receiving_tds", 0) * o["receiving_td"]
    pts += statline.get("receptions", 0)    * o["reception"]
    pts += statline.get("receiving_yds", 0) * y["receiving"]
    if statline.get("receiving_yds", 0) >= 100:
        pts += b["receiving_100_yds"]
    pts += statline.get("fumbles", 0)       * o["fumble_lost"]
    return pts

def simulate_team_score(team_stats: Dict[str, float]) -> int:
    td_count = (
        team_stats.get("passing_tds", 0) +
        team_stats.get("rushing_tds", 0) +
        team_stats.get("receiving_tds", 0)
    )
    pts = td_count * 6
    pts += sum(1 for _ in range(int(td_count)) if random.random() < 0.98)

    fg_attempts  = np.random.poisson(lam=FG_LAMBDA)
    fg_successes = sum(1 for _ in range(fg_attempts) if random.random() < 0.85)
    pts += fg_successes * 3

    two_pt_tries = np.random.binomial(int(td_count), 0.10)
    two_pt_made  = sum(1 for _ in range(two_pt_tries) if random.random() < 0.50)
    pts += two_pt_made * 2

    return pts

def simulate_for(
    matchups, teams, scoring, sims,
    espn_map, situational_map, def_ranks, vegas_map
):
    player_agg = {
        p["player"]: {c: [] for c in OUTPUT_CATEGORIES + ["dk_points"]}
        for roster in teams.values() for p in roster
    }
    team_agg = {t: {"wins":0,"scores":[],"spreads":[]} for t in teams}

    for away, home in matchups:
        key  = f"{away}-{home}"
        mods = situational_map.get(key, {})

        for _ in range(sims):

            def run_lineup(roster, is_home, mods, opponent, def_ranks):
                out = {}
                home_adv = mods.get("home_adv",0)
                weather  = mods.get("weather",0)
                fatigue  = mods.get(
                    "fatigue_home" if is_home else "fatigue_away",0.0
                )
                situat_base = (home_adv if is_home else -home_adv) + weather + fatigue

                # ── ESPN tilt
                espn = espn_map.get(key,{})
                home_wp = espn.get("home_win_prob")
                if home_wp is not None:
                    if home_wp > 1:
                        home_wp /= 100.0
                    situat_base += (home_wp-0.5)*2*ESPN_WEIGHT if is_home else -(home_wp-0.5)*2*ESPN_WEIGHT

                adj_pass, adj_rush, situ_mods = [], [], []
                for p in roster:
                    rank      = def_ranks.get(opponent, {}).get(f"vs_{p['position']}",16)
                    norm      = (16.5-rank)/16.5
                    def_bonus = norm*0.15
                    situ_p    = situat_base + def_bonus
                    situ_mods.append(situ_p)

                    adj_pass.append(p["stats"]["passing_yds"]*(1+situ_p))
                    adj_rush.append(p["stats"]["rushing_yds"]*(1+situ_p))

                sum_pass = sum(adj_pass); sum_rush = sum(adj_rush)
                team_pass = random.gauss(sum_pass,max(sum_pass*0.10,10))
                team_rush = random.gauss(sum_rush,max(sum_rush*0.15,5))

                sum_tgt  = sum(p["stats"]["receiving_targets"] for p in roster)
                team_tgt = random.gauss(sum_tgt,max(sum_tgt*0.15,5))

                for p, μp, μr, situ_p in zip(roster,adj_pass,adj_rush,situ_mods):
                    jitter = {}
                    share_p = μp/sum_pass if sum_pass>0 else 0
                    jitter["passing_yds"] = max(0,random.gauss(team_pass*share_p,max(μp*0.1,2)))
                    share_r = μr/sum_rush if sum_rush>0 else 0
                    jitter["rushing_yds"] = max(0,random.gauss(team_rush*share_r,max(μr*0.15,1)))

                    jitter["interceptions"] = sample_nb(p["stats"]["interceptions"])
                    jitter["fumbles"]       = sample_nb(p["stats"]["fumbles"])
                    jitter["passing_tds"]   = np.random.poisson(lam=p["stats"]["passing_tds"])
                    jitter["rushing_tds"]   = np.random.poisson(lam=p["stats"]["rushing_tds"])

                    base_td = p["stats"].get("receiving_tds",0)
                    raw_td  = base_td*(1+situ_p+0.1)
                    lam     = max(0.0, raw_td if not np.isnan(raw_td) else 0.0)
                    jitter["receiving_tds"] = np.random.poisson(lam=lam)

                    mu_ry     = p["stats"].get("receiving_yds",0)*(1+situ_p)
                    boosted   = mu_ry*1.1
                    sigma_ry  = max(mu_ry*0.15,2)
                    raw_ry    = random.gauss(boosted,sigma_ry)
                    jitter["receiving_yds"] = max(0,min(raw_ry,mu_ry*(CLAMP_MULT+0.5)))

                    share_t = p["stats"]["receiving_targets"]/sum_tgt if sum_tgt>0 else 0
                    raw_t   = team_tgt*share_t
                    jitter["receiving_targets"] = max(
                        0,
                        random.gauss(raw_t,max(p["stats"]["receiving_targets"]*0.1,1))
                    )

                    base_tg  = p["stats"].get("receiving_targets",1)
                    base_rec = p["stats"].get("receptions",0)
                    raw_rate = base_rec/base_tg if base_tg>0 else 0.65
                    catch_rate = min(max(raw_rate+0.05,0.0),1.0)
                    n_tgt = int(round(jitter["receiving_targets"]))
                    jitter["receptions"] = min(n_tgt,np.random.binomial(n_tgt,catch_rate))

                    for st in ["pass_attempts","pass_completions"]:
                        mu  = p["stats"].get(st,0)
                        raw = random.gauss(mu,max(mu*0.12,1))
                        jitter[st] = max(0,min(raw,mu*CLAMP_MULT))

                    dk_pts = score_statline(jitter,scoring)
                    out[p["player"]] = {"stats":jitter,"dk":dk_pts}

                return out

            h_out = run_lineup(teams[home],True, mods, away, def_ranks)
            a_out = run_lineup(teams[away],False,mods, home, def_ranks)

            h_stats = {c:sum(i["stats"].get(c,0) for i in h_out.values()) for c in OUTPUT_CATEGORIES}
            a_stats = {c:sum(i["stats"].get(c,0) for i in a_out.values()) for c in OUTPUT_CATEGORIES}

            h_score = round(simulate_team_score(h_stats))
            a_score = round(simulate_team_score(a_stats))

            # ── ESPN calibration: blend + noise ─────────────────────────────────
            espn      = espn_map.get(key,{})
            home_proj = espn.get("predicted_home_score") or espn.get("home_proj")
            away_proj = espn.get("predicted_away_score") or espn.get("away_proj")
            esp_ou    = espn.get("over_under", espn.get("total"))

            if home_proj is not None and away_proj is not None:
                h_cal = h_score*(1-ESPN_WEIGHT) + home_proj*ESPN_WEIGHT
                a_cal = a_score*(1-ESPN_WEIGHT) + away_proj*ESPN_WEIGHT
                sigma = max((esp_ou or 0)*0.05, 1)
                h_score = max(0, random.gauss(h_cal, sigma))
                a_score = max(0, random.gauss(a_cal, sigma))

            # ── dampen ─────────────────────────────────────────────────────────
            h_score = round(h_score * SCORE_DAMPEN)
            a_score = round(a_score * SCORE_DAMPEN)

            # ── Vegas calibration: blend + noise ────────────────────────────────
            vegas = vegas_map.get(key,{})
            mk_sp = vegas.get("spread")
            mk_ou = vegas.get("over_under", vegas.get("total"))
            if mk_sp is not None:
                total    = h_score + a_score
                diff_cal = (h_score - a_score)*0.5 + mk_sp*0.5
                sigma2   = max((mk_ou or 0)*0.02, 1)
                new_diff = random.gauss(diff_cal, sigma2)
                h_score  = max(0, (total + new_diff)/2)
                a_score  = max(0, (total - new_diff)/2)

            spread = h_score - a_score

            if h_score > a_score:
                team_agg[home]["wins"] += 1
            elif a_score > h_score:
                team_agg[away]["wins"] += 1
            else:
                team_agg[home]["wins"] += 0.5
                team_agg[away]["wins"] += 0.5

            for pl, info in {**h_out,**a_out}.items():
                for c,v in info["stats"].items():
                    player_agg[pl][c].append(v)
                player_agg[pl]["dk_points"].append(info["dk"])

            team_agg[home]["scores"].append(h_score)
            team_agg[away]["scores"].append(a_score)
            team_agg[home]["spreads"].append(spread)
            team_agg[away]["spreads"].append(-spread)

    return player_agg, team_agg

# ── SUMMARIES & ENTRYPOINT ─────────────────────────────────────────────────────

def summarize_players(
    player_agg: Dict[str, Dict[str, List[float]]],
    meta:       Dict[str, Dict[str, str]]
) -> Dict[str, Dict]:
    summary: Dict[str, Dict] = {}
    for pl, stats in player_agg.items():
        m   = meta.get(pl, {})
        rec = {
            "id":   m.get("id",   ""),
            "team": m.get("team", ""),
            "pos":  m.get("pos",  "")
        }
        for c, vals in stats.items():
            if not vals:
                continue
            rec[c] = {
                "floor":   min(vals),
                "mean":    statistics.mean(vals),
                "ceiling": max(vals)
            }
        rec["total_tds"] = (
            rec.get("rushing_tds",{}).get("mean",0) +
            rec.get("receiving_tds",{}).get("mean",0)
        )
        summary[pl] = rec
    return summary

def summarize_teams(
    team_agg: Dict[str, Dict[str, List[float]]],
    sims:     int
) -> Dict[str, Dict]:
    summary: Dict[str, Dict] = {}
    for t, agg in team_agg.items():
        if not agg["scores"]:
            summary[t] = {"win_prob":0.0, "mean_score":0.0, "mean_spread":0.0}
        else:
            summary[t] = {
                "win_prob":    agg["wins"]    / sims,
                "mean_score":  sum(agg["scores"])  / sims,
                "mean_spread": sum(agg["spreads"]) / sims
            }
    return summary

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-b","--baseline", type=Path, help="Baseline JSON")
    parser.add_argument("-s","--scoring",  type=Path, help="Scoring JSON")
    parser.add_argument("-w","--week",     type=int, default=1, help="Week number")
    parser.add_argument("-n","--sims",     type=int, default=1000, help="Number of sims")
    parser.add_argument("--year",          type=int, default=2025, help="Season year")
    parser.add_argument(
        "--days", nargs="+",
        choices=["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday","All"],
        default=["Sunday","Monday"],
        help="Weekdays to simulate; 'All'=every game"
    )
    parser.add_argument(
        "--matchup", action="append", nargs=2, metavar=("AWAY","HOME"),
        help="Filter specific games; e.g. --matchup KC PHI"
    )
    args = parser.parse_args()

    # ── Prepare file paths and load projection data ───────────────────────────────
    TRACKING_FILE   = PROJECT_ROOT / "src" / "data" / "tracking_results" / f"{args.year}_final_betting.xlsx"
    dvoa_map        = load_json(PROJECT_ROOT / "src" / "data" / "projection_data" / f"dvoa_{args.year}.json")
    espn_map        = load_json(ESPN_DIR  / f"espn_{args.year}_w{args.week}.json")  if (ESPN_DIR  / f"espn_{args.year}_w{args.week}.json").exists()  else {}
    vegas_map       = load_json(VEGAS_DIR  / f"vegas_{args.year}_w{args.week}.json") if (VEGAS_DIR  / f"vegas_{args.year}_w{args.week}.json").exists() else {}
    def_rank_map    = load_json(DEF_RANK_DIR / f"def_rank_{args.year}_w{args.week}.json") if (DEF_RANK_DIR / f"def_rank_{args.year}_w{args.week}.json").exists() else {}
    situational_map = load_json(SITUATION_DIR / f"week{args.week}.json")           if (SITUATION_DIR / f"week{args.week}.json").exists()           else {}
    team_profile    = load_json(TEAM_PROFILE) if TEAM_PROFILE.exists() else {}

    # ── Normalize ESPN win_prob entries stored as percentages ────────────────────
    for rec in espn_map.values():
        hp = rec.get("home_win_prob")
        ap = rec.get("away_win_prob")
        if hp is not None and hp > 1:
            rec["home_win_prob"] = hp / 100.0
        if ap is not None and ap > 1:
            rec["away_win_prob"] = ap / 100.0

    # ── Baseline & scoring inputs ────────────────────────────────────────────────
    baseline = args.baseline or (PROJS_ROOT / f"{args.year}_w{args.week}" / f"baseline_{args.year}_w{args.week}.json")
    scoring  = args.scoring  or SCORING_JSON
    if not baseline.exists():
        raise FileNotFoundError(f"Baseline not found: {baseline}")
    if not scoring.exists():
        raise FileNotFoundError(f"Scoring not found:  {scoring}")

    # ── Prepare players and schedule ─────────────────────────────────────────────
    players     = prepare_players(baseline, args.week)
    player_meta = {p["player"]:{"id":p["id"],"team":p["team"],"pos":p["position"]} for p in players}

    sched = nfl.import_schedules([args.year])
    sched["home_team"] = sched.home_team.map(lambda t: team_map.get(t,t))
    sched["away_team"] = sched.away_team.map(lambda t: team_map.get(t,t))
    week_sched = sched[sched["week"] == args.week]

    if args.matchup:
        mask = False
        for aw, hm in args.matchup:
            mask |= ((week_sched["away_team"] == aw) & (week_sched["home_team"] == hm))
        main_games = week_sched[mask]
    elif "All" in args.days:
        main_games = week_sched
    else:
        main_games = week_sched[week_sched["weekday"].isin(args.days)]

    main_matchups = [(r.away_team, r.home_team) for r in main_games.itertuples()]
    teams: Dict[str, List[Dict]] = {}
    for p in players:
        teams.setdefault(p["team"], []).append(p)
    for aw, hm in main_matchups:
        teams.setdefault(aw, [])
        teams.setdefault(hm, [])

    # ── Run simulations ───────────────────────────────────────────────────────────
    scoring_rules = load_json(scoring)["scoring"]
    player_agg, team_agg = simulate_for(
        main_matchups,
        teams,
        scoring_rules,
        args.sims,
        espn_map,
        situational_map,
        def_rank_map,
        vegas_map
    )

    # ── Write outputs ─────────────────────────────────────────────────────────────
    label    = args.days[0] if len(args.days) == 1 else "all"
    sims_dir = SIMS_ROOT / f"{args.year}_w{args.week}_{label}"
    sims_dir.mkdir(parents=True, exist_ok=True)

    PLAYER_SUMMARY = sims_dir / "player_summary.json"
    ps_main        = summarize_players(player_agg, player_meta)
    write_json(PLAYER_SUMMARY, ps_main)

    # Cleaned player summary Excel
    rows = []
    for name, rec in ps_main.items():
        row = {
            "player": name,
            "team": rec.get("team",""),
            "pos": rec.get("pos",""),
            "total_tds": round(rec.get("total_tds",0), 2),
            "total_yds_mean": round(
                rec.get("rushing_yds",{}).get("mean",0) +
                rec.get("receiving_yds",{}).get("mean",0), 2
            ),
            "pass_attempts_mean": round(rec.get("pass_attempts",{}).get("mean",0), 2),
            "completions_mean":   round(rec.get("pass_completions",{}).get("mean",0), 2),
            "passing_yds_mean":   round(rec.get("passing_yds",{}).get("mean",0), 2),
            "passing_yds_ceiling":round(rec.get("passing_yds",{}).get("ceiling",0), 2),
            "passing_tds_mean":   round(rec.get("passing_tds",{}).get("mean",0), 2),
            "interceptions_mean": round(rec.get("interceptions",{}).get("mean",0), 2),
            "rushing_yds_mean":   round(rec.get("rushing_yds",{}).get("mean",0), 2),
            "rushing_yds_ceiling":round(rec.get("rushing_yds",{}).get("ceiling",0), 2),
            "rushing_tds_mean":   round(rec.get("rushing_tds",{}).get("mean",0), 2),
            "receiving_targets_mean":   round(rec.get("receiving_targets",{}).get("mean",0),2),
            "receiving_targets_ceiling":round(rec.get("receiving_targets",{}).get("ceiling",0),2),
            "receptions_mean":     round(rec.get("receptions",{}).get("mean",0), 2),
            "receptions_ceiling":  round(rec.get("receptions",{}).get("ceiling",0), 2),
            "receiving_yds_mean":  round(rec.get("receiving_yds",{}).get("mean",0), 2),
            "receiving_yds_ceiling":round(rec.get("receiving_yds",{}).get("ceiling",0), 2),
            "receiving_tds_mean":  round(rec.get("receiving_tds",{}).get("mean",0), 2),
            "dk_points_mean":      round(rec.get("dk_points",{}).get("mean",0), 2),
            "dk_points_ceiling":   round(rec.get("dk_points",{}).get("ceiling",0), 2)
        }
        rows.append(row)

    pd.DataFrame(rows).to_excel(sims_dir / "player_summary.xlsx", index=False)
    print(f"Wrote {PLAYER_SUMMARY} & cleaned player_summary.xlsx")

    # Betting CSV & XLSX
    header = [
        "type","home","away",
        "home_win_prob","away_win_prob",
        "home_avg_score","away_avg_score",
        "mean_total","mean_spread"
    ]
    rows = []
    ts = summarize_teams(team_agg, args.sims)
    for r in main_games.itertuples():
        h, a = r.home_team, r.away_team
        hh, aa = ts[h], ts[a]
        hw, aw = hh["win_prob"], aa["win_prob"]
        total = hw + aw
        rows.append([
            "GAME", h, a,
            f"{(hw/total if total else 0.5):.3f}",
            f"{(aw/total if total else 0.5):.3f}",
            round(hh["mean_score"],1),
            round(aa["mean_score"],1),
            round(hh["mean_score"] + aa["mean_score"],1),
            round(hh["mean_score"] - aa["mean_score"],1)
        ])

    write_csv(sims_dir / "betting.csv", header, rows)
    pd.read_csv(sims_dir / "betting.csv").to_excel(sims_dir / "betting.xlsx", index=False)
    print(f"Wrote betting.csv & betting.xlsx to {sims_dir}")

# ── Append to tracking workbook using sheet copy as a template ───────────────
from copy import copy

if TRACKING_FILE.exists():
    wb        = load_workbook(TRACKING_FILE)
    template  = wb["template"]               # your formatted template sheet
    sheet_name = f"{args.year}_w{args.week}"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    new_ws = wb.copy_worksheet(template)
    new_ws.title = sheet_name

    df_bet = pd.read_csv(sims_dir / "betting.csv")

        # place each game row every 4 rows, starting at row 2
    for idx, data_row in enumerate(dataframe_to_rows(df_bet, index=False, header=False)):
        target_row = 2 + idx * 4
        for col_idx, val in enumerate(data_row, start=1):
            dest_cell = new_ws.cell(row=target_row, column=col_idx)
                
            # write the new value
            dest_cell.value = val

                # copy style from the same cell in the template sheet
            tmpl_cell      = template.cell(row=target_row, column=col_idx)
            dest_cell.fill = copy(tmpl_cell.fill)
            dest_cell.font = copy(tmpl_cell.font)
            dest_cell.border = copy(tmpl_cell.border)
            dest_cell.alignment = copy(tmpl_cell.alignment)
            dest_cell.number_format = tmpl_cell.number_format

        wb.save(TRACKING_FILE)
        print(f"Appended week {args.week} to {TRACKING_FILE} with template styling preserved")